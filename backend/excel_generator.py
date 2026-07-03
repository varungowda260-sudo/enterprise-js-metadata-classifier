"""
Excel Report Generator for JavaScript Metadata Classification.

Generates multi-sheet Excel reports with streaming for large datasets.
"""
import io
from datetime import datetime
from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from models import MetadataRecord, ClassificationResult, SkippedFile, ProcessingStats


class ExcelGenerator:
    """
    Generates Excel reports with classification results.
    Uses streaming for memory-efficient large dataset handling.
    """

    def __init__(self):
        self.workbook = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Setup cell styles for consistent formatting."""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.number_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self.error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    def create_summary_sheet(
        self,
        classifications: List[ClassificationResult]
    ) -> None:
        """
        Create Summary sheet with sys_name, total records, unique count, and note_unids.

        Args:
            classifications: List of classification results
        """
        ws = self.workbook.active
        ws.title = "Summary"

        # Headers
        headers = ["Sys Name", "Total Valid Records", "Unique Note Count", "Note UNIDs"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # Data rows
        for row, result in enumerate(classifications, 2):
            ws.cell(row=row, column=1, value=result.sys_name).border = self.border
            ws.cell(row=row, column=2, value=result.valid_records).border = self.border
            ws.cell(row=row, column=3, value=result.unique_note_count).border = self.border
            cell=ws.cell(row=row, column=4, value="\n".join(result.note_unids))
            cell.border = self.border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = max(18, len(result.note_unids) * 18)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 45

    def create_details_sheet(
        self,
        valid_records: List[MetadataRecord]
    ) -> None:
        """
        Create Details sheet with file info.

        Args:
            valid_records: List of valid metadata records
        """
        ws = self.workbook.create_sheet("Details")

        # Headers
        headers = ["File Name", "Sys Name", "Change Request Status","Change Request No.","Go Live Date Production","Change Request Closure Date","Note UNID"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = self.border

        # Data rows - stream in chunks for memory efficiency
        for row, record in enumerate(valid_records, 2):
            closure_dates = []
            if record.itqm_closure_date:
                closure_dates.append(record.itqm_closure_date)
            if record.cqa_closure_date:
                closure_dates.append(record.cqa_closure_date)
                
            closure_value = "\n".join(closure_dates)  
            ws.cell(row=row, column=1, value=record.file_name).border = self.border
            ws.cell(row=row, column=2, value=record.sys_name).border = self.border
            ws.cell(row=row, column=3, value=record.status).border = self.border
            ws.cell(row=row, column=4, value=record.cc_number).border = self.border
            ws.cell(row=row, column=5, value=record.implementation_date).border = self.border
            cell = ws.cell(row=row, column=6, value=closure_value)
            cell.border = self.border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if len(closure_dates) > 1:
                ws.row_dimensions[row].height = 36
            cell = ws.cell(row=row, column=7, value=record.note_unid)
            cell.border = self.border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 28
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 40

    def create_skipped_sheet(
        self,
        skipped_files: List[SkippedFile]
    ) -> None:
        """
        Create Skipped Files sheet.

        Args:
            skipped_files: List of skipped file records
        """
        ws = self.workbook.create_sheet("Skipped Files")

        # Headers
        headers = ["File Name", "Reason Skipped"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # Data rows
        for row, skipped in enumerate(skipped_files, 2):
            ws.cell(row=row, column=1, value=skipped.file_name).border = self.border
            ws.cell(row=row, column=2, value=skipped.reason).border = self.border
            cell = ws.cell(row=row, column=2)
            cell.fill = self.error_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 60

    def create_statistics_sheet(
        self,
        stats: ProcessingStats,
        duplicates_removed: int = 0
    ) -> None:
        """
        Create Processing Statistics sheet.

        Args:
            stats: Processing statistics
            duplicates_removed: Count of duplicate records removed
        """
        ws = self.workbook.create_sheet("Processing Statistics")

        # Headers
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # Data rows
        metrics = [
            ("Processing Time", f"{stats.elapsed_seconds:.2f} seconds"),
            ("Files Per Second", f"{stats.files_per_second:.2f}"),
            ("Total Files", stats.total_files_found),
            ("Valid Files", stats.files_processed - stats.files_with_errors),
            ("Skipped Files", stats.files_with_errors),
            ("Cancelled Records Ignored", stats.cancelled_records),
            ("Duplicate Records Removed", duplicates_removed),
            ("Unique Systems", stats.unique_systems),
            ("Status", stats.status.value),
        ]

        for row, (metric, value) in enumerate(metrics, 2):
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

    def generate(
        self,
        classifications: List[ClassificationResult],
        valid_records: List[MetadataRecord],
        skipped_files: List[SkippedFile],
        stats: ProcessingStats,
        duplicates_removed: int = 0
    ) -> bytes:
        """
        Generate complete Excel report.

        Returns:
            Bytes of the Excel file
        """
        self.__init__()  # Reset workbook

        self.create_summary_sheet(classifications)
        self.create_details_sheet(valid_records)

        if skipped_files:
            self.create_skipped_sheet(skipped_files)

        self.create_statistics_sheet(stats, duplicates_removed)

        # Write to bytes buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def save_to_file(
        self,
        file_path: Path,
        classifications: List[ClassificationResult],
        valid_records: List[MetadataRecord],
        skipped_files: List[SkippedFile],
        stats: ProcessingStats,
        duplicates_removed: int = 0
    ) -> None:
        """Save Excel report to file."""
        excel_bytes = self.generate(
            classifications, valid_records, skipped_files, stats, duplicates_removed
        )
        Path(file_path).write_bytes(excel_bytes)
